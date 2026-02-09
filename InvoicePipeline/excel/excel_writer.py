from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from suppliers.lenovo_config import COLUMN_MAPPING


# ==========================================================
# COMMON HELPERS
# ==========================================================

def auto_adjust_width(sheet, max_width=50):
    for col_idx, col in enumerate(sheet.columns, start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        sheet.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def style_header(cell):
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", start_color="E8E8E8")
    cell.alignment = Alignment(wrap_text=True, horizontal="center")


def style_title(cell):
    cell.font = Font(bold=True, size=14)
    cell.fill = PatternFill("solid", start_color="D3D3D3")
    cell.alignment = Alignment(horizontal="center")


# ==========================================================
# LENOVO EXCEL WRITER (LINE LEVEL)
# ==========================================================
def write_lenovo_excel(rows, output_path):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "View Supplier Invoice"

    headers = list(COLUMN_MAPPING.keys())

    # Title
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = sheet.cell(row=1, column=1)
    title.value = "Invoice Lines"
    title.font = Font(bold=True, size=14)
    title.alignment = Alignment(horizontal="center")
    title.fill = PatternFill("solid", start_color="D3D3D3")

    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="E8E8E8")
        cell.alignment = Alignment(wrap_text=True)

    # Data rows
    for row_idx, data in enumerate(rows, 3):
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = data.get(header, "")
            cell.value = value
            cell.alignment = Alignment(wrap_text=True)

            if header in ["Unit Cost", "Extended Amount", "Tax"]:
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0.00"

    # Auto width
    for col_idx, col in enumerate(sheet.columns, start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = min(max_len + 2, 50)

    sheet.freeze_panes = "A3"
    wb.save(output_path)


# ==========================================================
# MEYER EXCEL WRITER (LOCATION SUMMARY)
# ==========================================================

def write_meyer_excel(location_totals, location_meta, json_grand_total, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Headers
    ws.append([
        "Location Name & Address",
        "Location Identifier",
        "Address",
        "Postal Code",
        "Total Amount"
    ])

    # Data rows
    for location, total in location_totals.items():
        meta = location_meta.get(location, {})
        ws.append([
            location,
            meta.get("location_identifier", ""),
            meta.get("address", ""),
            meta.get("postal_code", ""),
            round(total, 2)
        ])

    # Empty row
    ws.append(["", "", "", "", ""])

    # Grand total
    ws.append(["GRAND TOTAL", "", "", "", round(json_grand_total, 2)])

    wb.save(output_path)
