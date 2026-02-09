import json
import re
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from config import OUTPUT_DIR

# ==========================================================
# CONFIGURATION
# ==========================================================

COLUMN_MAPPING = {
    "Invoice Number": "invoice.metadata.invoiceNumber",   # ✅ NEW
    "Invoice Line": None,
    "Company": "customer.name",
    "Item": None,
    "Line Item Description": {
        "template": "{description}, {serial} WK {week}"
    },
    "Commodity Code": None,
    "Spend Category": None,
    "Ship-To Address": "customer.address",
    "Ship-To Contact": None,
    "Tax": "line_item.vatPercentage",
    "Tax Recoverability": None,
    "Tax Option": None,
    "Quantity": {"default": 1},
    "Unit of Measure": {"default": "EA"},
    "Unit Cost": "line_item.unitPrice",
    "Extended Amount": "line_item.unitPrice",
    "Item Identifiers": None,
    "Memo": {
        "template": "{description}, {serial} WK {week}"
    },
    "Cost Center": "line_item.costCenter",
    "Location": None,
    "Intercompany Affiliate": None,
    "Inbound Streams": None,
    "Additional Worktags": None,
    "Worktag Split Template": None,
    "Split Button Count": {"default": 0},
    "Splits": {"default": 0},
}

# ==========================================================
# UTILITIES
# ==========================================================

def extract_serial_numbers(serial_string):
    if not serial_string:
        return []
    return [s.strip() for s in re.split(r"\s+", serial_string) if s.strip()]

def get_value_from_path(obj, path):
    if not path:
        return ""
    for part in path.split("."):
        if not isinstance(obj, dict):
            return ""
        obj = obj.get(part, "")
    return obj

def format_template(template, context):
    try:
        return template.format(**context)
    except KeyError:
        return template

def get_week_from_invoice_date(invoice_date_str):
    if not invoice_date_str:
        return ""

    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            invoice_date = datetime.strptime(invoice_date_str, fmt).date()
            return invoice_date.isocalendar()[1]
        except ValueError:
            continue

    return ""

# ==========================================================
# CORE PARSER
# ==========================================================

def parse_json_to_rows(json_data):
    rows = []
    seen = set()

    records = json_data if isinstance(json_data, list) else [json_data]

    for record in records:
        data_blocks = record.get("result", {}).get("data", [])

        for block in data_blocks:
            invoice = block.get("extracted_data", {}).get("invoice", {})
            if not invoice:
                continue

            metadata = invoice.get("metadata", {})
            customer = invoice.get("customer", {})
            line_items = invoice.get("lineItems", [])

            invoice_number = metadata.get("invoiceNumber", "")
            invoice_date = metadata.get("invoiceDate", "")
            week_number = get_week_from_invoice_date(invoice_date)

            for line_item in line_items:
                serials = extract_serial_numbers(
                    line_item.get("serialNumbers", "")
                ) or [""]

                for serial in serials:
                    unique_key = (
                        invoice_number,
                        line_item.get("orderLineNumber", ""),
                        serial
                    )

                    if unique_key in seen:
                        continue
                    seen.add(unique_key)

                    context = {
                        "description": line_item.get("description", ""),
                        "serial": serial,
                        "week": week_number,
                        "customer": customer,
                        "line_item": line_item,
                        "invoice": {"metadata": metadata},  # ✅ important
                    }

                    row = {}

                    for col, rule in COLUMN_MAPPING.items():
                        if rule is None:
                            row[col] = ""
                            continue

                        if isinstance(rule, dict) and "default" in rule:
                            row[col] = rule["default"]
                            continue

                        if isinstance(rule, dict) and "template" in rule:
                            row[col] = format_template(rule["template"], context)
                            continue

                        if isinstance(rule, str):
                            row[col] = get_value_from_path(context, rule)

                    rows.append(row)

    return rows

# ==========================================================
# EXCEL CREATION
# ==========================================================

def create_excel(rows, output_path):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "View Supplier Invoice"

    headers = list(COLUMN_MAPPING.keys())

    # Title
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = sheet.cell(row=1, column=1)
    title.value = "Invoice Lines"
    title.font = Font(bold=True, size=14)
    title.alignment = Alignment(horizontal="center")
    title.fill = PatternFill("solid", start_color="D3D3D3")

    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="E8E8E8")
        cell.alignment = Alignment(wrap_text=True)

    # Data rows
    for row_idx, data in enumerate(rows, 3):
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = data.get(header, "")
            cell.value = value
            cell.alignment = Alignment(wrap_text=True)

            if header in ["Unit Cost", "Extended Amount", "Tax"]:
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0.00"

    # Auto column width
    for col_idx, col in enumerate(sheet.columns, start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = min(max_len + 2, 50)

    sheet.freeze_panes = "A3"
    wb.save(output_path)

# ==========================================================
# MAIN
# ==========================================================

def main():
    if len(sys.argv) != 2:
        print("Usage: python lenovo_report.py <json_file_path>")
        sys.exit(1)

    json_path = Path(sys.argv[1])

    if not json_path.exists():
        print(f"❌ JSON file not found: {json_path}")
        sys.exit(1)

    try:
        with open(json_path, "r", encoding="utf-8") as f:
            json_data = json.load(f)

        rows = parse_json_to_rows(json_data)

        if not rows:
            print("❌ No rows extracted")
            sys.exit(1)

        data_block = json_data.get("result", {}).get("data", [{}])[0]
        original_filename = data_block.get("filename", json_path.stem)
        pdf_name = Path(original_filename).stem
        output_file = OUTPUT_DIR / f"{pdf_name}_lenovo_report.xlsx"

        create_excel(rows, output_file)

        print(f"✅ Excel created: {output_file}")
        print(f"🧾 Rows written: {len(rows)}")

    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
