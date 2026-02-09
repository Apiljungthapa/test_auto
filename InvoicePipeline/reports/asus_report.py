import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from config import OUTPUT_DIR

# ==========================================================
# CONFIG
# ==========================================================

OUTPUT_HEADERS = [
    "id",
    "organization",
    "memo",
    "amount",
    "tax_rate",
    "tax_amount",
    "gross_amount"
]

# ==========================================================
# FUNCTIONS
# ==========================================================

def load_json(json_path: Path) -> dict:
    """Load JSON file from path"""
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def extract_rows(data: dict):
    """
    Extract valid line items based on rules:
    - Must have Lease_driver_name AND contract_number
    """
    rows = []

    documents = data.get("result", {}).get("data", [])

    for doc in documents:
        extracted = doc.get("extracted_data", {})
        line_items = extracted.get("line_items", [])

        for item in line_items:
            lease_driver = item.get("Lease_driver_name")
            contract_no = item.get("contract_number")

            # ❌ Skip if required fields are missing
            if not lease_driver or not contract_no:
                continue

            row = {
                "id": None,
                "organization": item.get("cost_center"),
                "memo": f"{item.get('Period', '')} - {lease_driver}",
                "amount": item.get("net_amount"),
                "tax_rate": item.get("tax_rate"),
                "tax_amount": item.get("tax_amount"),
                "gross_amount": item.get("gross_amount"),
            }

            rows.append(row)

    return rows


def write_excel(rows, output_path: Path):
    """Write extracted rows to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lease_Invoices"

    # Header row
    ws.append(OUTPUT_HEADERS)

    for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Data rows
    for row in rows:
        ws.append([row.get(col) for col in OUTPUT_HEADERS])

    wb.save(output_path)


# ==========================================================
# MAIN
# ==========================================================

def main():
    if len(sys.argv) < 2:
        print("Usage: python lease_json_to_excel.py <json_file_path>")
        sys.exit(1)

    json_path = Path(sys.argv[1])

    if not json_path.exists():
        print(f"JSON file not found: {json_path}")
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    data = load_json(json_path)
    rows = extract_rows(data)

    output_file = OUTPUT_DIR / f"{json_path.stem}_processed.xlsx"
    write_excel(rows, output_file)

    print(f"Excel file created successfully: {output_file}")


if __name__ == "__main__":
    main()
