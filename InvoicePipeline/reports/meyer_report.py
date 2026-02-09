import json
import sys
import re
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from config import OUTPUT_DIR

# ==========================================================
# CONFIG
# ==========================================================

REFERENCE_EXCEL_PATH = Path(
    "C:\\Users\\HP\\Desktop\\Invoice_pipeline_setup\\InvoicePipeline\\Location_Code\\Location & address.xlsx"
)

# ==========================================================
# UTILITIES
# ==========================================================

def load_json(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)

def extract_postal_code(text):
    """
    Extract postal code from address text
    Example matches: 560001, 90210
    """
    if not text:
        return ""
    match = re.search(r"\b\d{5,6}\b", text)
    return match.group(0) if match else ""

def load_reference_excel(excel_path):
    """
    Load reference Excel into dict:
    { postal_code: {location_id, address, postal_code} }
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    idx_location_id = headers.index("Location Identifier")
    idx_address = headers.index("Address")
    idx_postal = headers.index("Postal Code")

    ref_data = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        postal_code = str(row[idx_postal]).strip()
        if postal_code:
            ref_data[postal_code] = {
                "location_id": row[idx_location_id],
                "address": row[idx_address],
                "postal_code": postal_code
            }

    return ref_data

# ==========================================================
# CORE LOGIC
# ==========================================================

def aggregate_by_location(service_locations, reference_data):
    """
    Aggregate totals and enrich with reference Excel data
    """
    location_totals = defaultdict(float)
    location_meta = {}

    for location in service_locations:
        location_text = location.get("location_name_address", "").strip()
        postal_code = extract_postal_code(location_text)

        for item in location.get("line_items", []):
            try:
                location_totals[location_text] += float(item.get("total_line_price", 0))
            except (ValueError, TypeError):
                pass

        ref = reference_data.get(postal_code, {})

        location_meta[location_text] = {
            "location_identifier": ref.get("location_id", ""),
            "address": ref.get("address", ""),
            "postal_code": postal_code
        }

    return location_totals, location_meta

# ==========================================================
# EXCEL CREATION
# ==========================================================

def create_excel(location_totals, location_meta, json_grand_total, invoice_number, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Headers
    ws.append([
        "Invoice Number",
        "Location Name & Address",
        "Location Identifier",
        "Address",
        "Postal Code",
        "Total Amount"
    ])

    # Data rows
    for location, total in location_totals.items():
        meta = location_meta.get(location, {})
        ws.append([
            invoice_number,                         # repeated value
            location,
            meta.get("location_identifier", ""),
            meta.get("address", ""),
            meta.get("postal_code", ""),
            round(total, 2)
        ])

    # Empty row
    ws.append(["", "", "", "", "", ""])

    # Grand total row
    ws.append([
        "GRAND TOTAL",
        "",
        "",
        "",
        "",
        round(json_grand_total, 2)
    ])

    wb.save(output_path)

# ==========================================================
# MAIN
# ==========================================================

def main():
    if len(sys.argv) != 2:
        print("Usage: python meyer_report.py <json_file_path>")
        sys.exit(1)

    json_path = Path(sys.argv[1])

    if not json_path.exists():
        print(f"❌ JSON file not found: {json_path}")
        sys.exit(1)

    if not REFERENCE_EXCEL_PATH.exists():
        print(f"❌ Reference Excel not found: {REFERENCE_EXCEL_PATH}")
        sys.exit(1)

    try:
        print(f"📊 Loading JSON: {json_path.name}")
        data = load_json(json_path)

        print("📘 Loading reference Excel...")
        reference_data = load_reference_excel(REFERENCE_EXCEL_PATH)

        data_block = data.get("result", {}).get("data", [{}])[0]
        original_filename = data_block.get("filename", json_path.stem)
        pdf_name = Path(original_filename).stem

        extracted_data = data_block.get("extracted_data", {})

        if "invoice" in extracted_data:
            root = extracted_data.get("invoice", {})
        else:
            root = extracted_data

        service_locations = root.get("service_locations", [])

        if not service_locations:
            print("❌ No service locations found")
            sys.exit(1)

        # 🔹 Extract Invoice Number (same for all rows)
        invoice_number = service_locations[0].get("Invoice_number", "")

        json_grand_total = root.get("grand_total")

        location_totals, location_meta = aggregate_by_location(
            service_locations, reference_data
        )

        if json_grand_total is None:
            json_grand_total = sum(location_totals.values())

        output_path = OUTPUT_DIR / f"{pdf_name}_meyer_report.xlsx"

        print("📋 Creating Excel report...")
        create_excel(
            location_totals,
            location_meta,
            json_grand_total,
            invoice_number,
            output_path
        )

        print(f"✅ Excel created: {output_path}")
        print(f"📍 Locations processed: {len(location_totals)}")
        print(f"🧾 Invoice Number: {invoice_number}")

    except Exception as e:
        print(f"\n❌ Error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
